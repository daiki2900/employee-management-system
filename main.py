# main.py
import re
import tkinter as tk
from tkinter import messagebox, filedialog
from tkinter import ttk
import ttkbootstrap as ttkb
from db_setup import insert_employee, fetch_all_employees, update_employee, delete_employee
import csv
from fpdf import FPDF
import openpyxl
from openpyxl import Workbook


def export_to_excel():
    """Export employee data to an Excel file."""
    # Open a save file dialog to allow the user to specify the file name and location
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])

    # If the user cancels the dialog, return early
    if not file_path:
        return

    # Create a new Excel workbook and select the active worksheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Employee Data"

    # Add the headers to the first row of the worksheet
    headers = ["ID", "First Name", "Last Name",
               "Department", "Salary", "Hire Date"]
    sheet.append(headers)

    # Extract employee data from the Treeview and add it to the worksheet
    for record in employee_table.get_children():
        employee = employee_table.item(record)['values']
        sheet.append(employee)

    # Save the workbook to the specified file path
    workbook.save(file_path)

    # Show a success message
    messagebox.showinfo("Export Successful",
                        "Data has been exported to Excel file successfully!")


def generate_pdf_report():
    """Generate a PDF report of the employee data."""
    # Open a save file dialog to allow the user to specify the file name and location
    file_path = filedialog.asksaveasfilename(
        defaultextension=".pdf", filetypes=[("PDF files", "*.pdf")])

    # If the user cancels the dialog, return early
    if not file_path:
        return

    # Create a PDF object
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Set title for the PDF
    pdf.set_font("Arial", 'B', 16)
    pdf.cell(200, 10, txt="Employee Management Report", ln=True, align="C")

    # Leave some space
    pdf.ln(10)

    # Set header for the table
    pdf.set_font("Arial", 'B', 12)
    pdf.cell(20, 10, "ID", 1)
    pdf.cell(40, 10, "First Name", 1)
    pdf.cell(40, 10, "Last Name", 1)
    pdf.cell(40, 10, "Department", 1)
    pdf.cell(30, 10, "Salary", 1)
    pdf.cell(30, 10, "Hire Date", 1)
    pdf.ln()

    # Extract employee data from the Treeview
    pdf.set_font("Arial", '', 12)
    for record in employee_table.get_children():
        employee = employee_table.item(record)['values']
        pdf.cell(20, 10, str(employee[0]), 1)
        pdf.cell(40, 10, employee[1], 1)
        pdf.cell(40, 10, employee[2], 1)
        pdf.cell(40, 10, employee[3], 1)
        pdf.cell(30, 10, str(employee[4]), 1)
        pdf.cell(30, 10, employee[5], 1)
        pdf.ln()

    # Save the PDF to the specified location
    pdf.output(file_path)

    # Show a success message
    messagebox.showinfo("Report Generated",
                        "PDF report has been generated successfully!")


def export_to_csv():
    """Export employee data to a CSV file."""
    # Open a save file dialog to allow the user to specify the file name and location
    file_path = filedialog.asksaveasfilename(
        defaultextension=".csv", filetypes=[("CSV files", "*.csv")])

    # If the user cancels the dialog, return early
    if not file_path:
        return

    # Extract the employee data from the Treeview
    employee_data = []
    for record in employee_table.get_children():
        employee_data.append(employee_table.item(record)['values'])

    # Write the data to the CSV file
    with open(file_path, mode="w", newline="") as file:
        writer = csv.writer(file)
        # Write the header (column names)
        writer.writerow(["ID", "First Name", "Last Name",
                        "Department", "Salary", "Hire Date"])
        # Write the employee data
        writer.writerows(employee_data)

    # Show a success message
    messagebox.showinfo(
        "Export Successful", "Employee data has been exported to the CSV file successfully!")


def search_employees():
    """Search for employees by first or last name and display the results."""
    search_term = search_entry.get().lower(
    )  # Get the search term and convert to lowercase

    # Clear the current display in the Treeview
    for record in employee_table.get_children():
        employee_table.delete(record)

    # Fetch all employees and filter by search term
    employees = fetch_all_employees()
    filtered_employees = [employee for employee in employees if search_term in employee[1].lower(
    ) or search_term in employee[2].lower()]

    # Insert the filtered data into the Treeview
    for employee in filtered_employees:
        employee_table.insert("", "end", values=employee)


def validate_inputs(first_name, last_name, department, salary, hire_date):
    """Validates the form inputs and returns True if valid, False otherwise."""

    # Check if any field is empty
    if not all([first_name, last_name, department, salary, hire_date]):
        messagebox.showerror("Input Error", "All fields are required.")
        return False

    # Validate salary (must be a positive number)
    try:
        salary = float(salary)
        if salary <= 0:
            messagebox.showerror(
                "Input Error", "Salary must be a positive number.")
            return False
    except ValueError:
        messagebox.showerror("Input Error", "Salary must be a valid number.")
        return False

    # Validate hire date (format: YYYY-MM-DD)
    if not re.match(r"\d{4}-\d{2}-\d{2}", hire_date):
        messagebox.showerror(
            "Input Error", "Hire date must be in the format YYYY-MM-DD.")
        return False

    return True


def load_employees():
    """Fetch all employee records and display them in the Treeview."""
    for record in employee_table.get_children():  # Clear existing data
        employee_table.delete(record)

    employees = fetch_all_employees()  # Fetch from the database

    # Insert fetched data into the Treeview
    for employee in employees:
        employee_table.insert("", "end", values=employee)


def add_employee():
    """Insert a new employee into the database and refresh the Treeview."""
    first_name = first_name_entry.get()
    last_name = last_name_entry.get()
    department = department_entry.get()
    salary = salary_entry.get()
    hire_date = hire_date_entry.get()

    # Validate inputs before inserting into the database
    if validate_inputs(first_name, last_name, department, salary, hire_date):
        insert_employee(first_name, last_name, department, salary, hire_date)
        load_employees()  # Refresh the Treeview
        messagebox.showinfo("Success", "Employee added successfully.")


def select_employee(event):
    """Populate form fields when an employee record is selected from the Treeview."""
    selected = employee_table.focus()
    values = employee_table.item(selected, 'values')

    if values:
        first_name_entry.delete(0, tk.END)
        last_name_entry.delete(0, tk.END)
        department_entry.delete(0, tk.END)
        salary_entry.delete(0, tk.END)
        hire_date_entry.delete(0, tk.END)

        first_name_entry.insert(0, values[1])
        last_name_entry.insert(0, values[2])
        department_entry.insert(0, values[3])
        salary_entry.insert(0, values[4])
        hire_date_entry.insert(0, values[5])


def update_employee_record():
    """Update an existing employee record in the database."""
    selected = employee_table.focus()
    values = employee_table.item(selected, 'values')

    if values:  # Ensure a record is selected
        employee_id = values[0]
        first_name = first_name_entry.get()
        last_name = last_name_entry.get()
        department = department_entry.get()
        salary = salary_entry.get()
        hire_date = hire_date_entry.get()

        # Validate inputs before updating the database
        if validate_inputs(first_name, last_name, department, salary, hire_date):
            update_employee(employee_id, first_name, last_name,
                            department, salary, hire_date)
            load_employees()  # Refresh the Treeview
            messagebox.showinfo("Success", "Employee updated successfully.")


def delete_employee_record():
    """Delete an employee record from the database."""
    selected = employee_table.focus()
    values = employee_table.item(selected, 'values')

    if values:  # Ensure a record is selected
        employee_id = values[0]
        delete_employee(employee_id)
        load_employees()  # Refresh the Treeview


# Create the main application window
# 'darkly' is one of ttkbootstrap's themes
app = ttkb.Window(themename="darkly")
app.title("Employee Management System")
app.geometry("900x600")  # Set the window size

# Frame for employee details form
form_frame = ttk.Frame(app)
form_frame.grid(row=0, column=0, padx=20, pady=20, sticky="nsew")

# Labels and Entry fields for employee details inside the form frame
ttk.Label(form_frame, text="First Name").grid(
    row=0, column=0, padx=10, pady=10, sticky="w")
first_name_entry = ttk.Entry(form_frame)
first_name_entry.grid(row=0, column=1, padx=10, pady=10)

ttk.Label(form_frame, text="Last Name").grid(
    row=1, column=0, padx=10, pady=10, sticky="w")
last_name_entry = ttk.Entry(form_frame)
last_name_entry.grid(row=1, column=1, padx=10, pady=10)

ttk.Label(form_frame, text="Department").grid(
    row=2, column=0, padx=10, pady=10, sticky="w")
department_entry = ttk.Entry(form_frame)
department_entry.grid(row=2, column=1, padx=10, pady=10)

ttk.Label(form_frame, text="Salary").grid(
    row=3, column=0, padx=10, pady=10, sticky="w")
salary_entry = ttk.Entry(form_frame)
salary_entry.grid(row=3, column=1, padx=10, pady=10)

ttk.Label(form_frame, text="Hire Date (YYYY-MM-DD)").grid(row=4,
                                                          column=0, padx=10, pady=10, sticky="w")
hire_date_entry = ttk.Entry(form_frame)
hire_date_entry.grid(row=4, column=1, padx=10, pady=10)


# Frame for buttons
button_frame = ttk.Frame(app)
button_frame.grid(row=1, column=0, padx=20, pady=10, sticky="ew")

# Buttons for Add, Update, and Delete, styled with ttkbootstrap themes
ttk.Button(button_frame, text="Add Employee", command=lambda: add_employee(
), bootstyle="success").grid(row=0, column=0, padx=10, pady=10)
ttk.Button(button_frame, text="Update Employee", command=lambda: update_employee_record(
), bootstyle="info").grid(row=0, column=1, padx=10, pady=10)
ttk.Button(button_frame, text="Delete Employee", command=lambda: delete_employee_record(
), bootstyle="danger").grid(row=0, column=2, padx=10, pady=10)

# Add the Export to CSV button in the button frame
ttk.Button(button_frame, text="Export to CSV", command=lambda: export_to_csv(
), bootstyle="info").grid(row=0, column=3, padx=10, pady=10)

# Add the Export to Excel button in the button frame
ttk.Button(button_frame, text="Export to Excel", command=lambda: export_to_excel(
), bootstyle="info").grid(row=0, column=4, padx=10, pady=10)


# Add the Generate PDF Report button in the button frame
ttk.Button(button_frame, text="Generate PDF Report", command=lambda: generate_pdf_report(
), bootstyle="info").grid(row=0, column=5, padx=10, pady=10)


# Frame for the search bar
search_frame = ttk.Frame(app)
search_frame.grid(row=3, column=0, padx=20, pady=10, sticky="ew")

# Search label and entry
ttk.Label(search_frame, text="Search:").grid(
    row=0, column=0, padx=10, pady=10, sticky="w")
search_entry = ttk.Entry(search_frame)
search_entry.grid(row=0, column=1, padx=10, pady=10, sticky="ew")

# Search button
ttk.Button(search_frame, text="Search", command=lambda: search_employees(
), bootstyle="primary").grid(row=0, column=2, padx=10, pady=10)


# Frame for the employee table (Treeview)
table_frame = ttk.Frame(app)
table_frame.grid(row=4, column=0, padx=20, pady=20, sticky="nsew")

# Treeview to display employee records inside the table frame
columns = ("ID", "First Name", "Last Name",
           "Department", "Salary", "Hire Date")
employee_table = ttk.Treeview(
    table_frame, columns=columns, show="headings", bootstyle="primary")

# Define column headers and their width
for col in columns:
    employee_table.heading(col, text=col)
    employee_table.column(col, width=120)

employee_table.grid(row=0, column=0, sticky="nsew")

# Scrollbar for the Treeview
scrollbar = ttk.Scrollbar(table_frame, orient="vertical",
                          command=employee_table.yview)
employee_table.configure(yscroll=scrollbar.set)
scrollbar.grid(row=0, column=1, sticky="ns")


# Triggered when a row is clicked
employee_table.bind("<ButtonRelease-1>", select_employee)

# Load employees on startup
load_employees()


# Run the application
app.mainloop()
